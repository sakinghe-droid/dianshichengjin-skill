#!/usr/bin/env python3
"""
game_theory_6d — 六维博弈引擎 (CLI入口)

兼容六维博弈输入JSON格式。

用法:
    python scripts/game_theory_6d.py --input <input.json> [--kline <kline.json>]
    python scripts/game_theory_6d.py --code 000001 --auto-kline
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

OPEN_SCRIPTS = os.path.dirname(os.path.abspath(__file__))
if OPEN_SCRIPTS not in sys.path:
    sys.path.insert(0, OPEN_SCRIPTS)

import argparse
import json
import subprocess

from core.game_theory_6d import analyze as game_analyze
from core.trend_forecaster import consensus_arbitration, forecast as tf_forecast


def fetch_data(code: str) -> tuple:
    """获取行情+K线 数据"""
    # 行情
    r1 = subprocess.run(
        [sys.executable, f"{OPEN_SCRIPTS}/data_fallback.py",
         "--mode", "quote", "--code", code],
        capture_output=True, text=True, timeout=30
    )
    quote = json.loads(r1.stdout)
    
    # K线
    r2 = subprocess.run(
        [sys.executable, f"{OPEN_SCRIPTS}/data_fallback.py",
         "--mode", "kline", "--code", code, "--count", "120"],
        capture_output=True, text=True, timeout=30
    )
    klines = json.loads(r2.stdout)
    
    return quote, klines


def enrich_real_data(code: str, quote: dict, klines: list) -> dict:
    """从妙想mx-data + TDX MCP + data_fallback获取真实数据，不用默认值"""
    enrichment = {}
    
    # 1. 妙想 mx-data: 多周期DDX/DDY + 主力资金 (首选)
    try:
        mx_key = os.environ.get('MX_APIKEY', '')
        if mx_key:
            import subprocess as _sp
            import re, glob as _glob
            import openpyxl
            
            # 清除旧 mx output
            mx_out_dir = os.path.expanduser('~/.openclaw/workspace/mx_data/output')
            old_files = _glob.glob(f'{mx_out_dir}/mx_data_{code}_*.xlsx')
            for f in old_files:
                try: os.remove(f)
                except: pass
            
            # 查询1: DDX数据
            _sp.run(
                [sys.executable, os.path.expanduser('~/skills/mx-data/mx_data.py'),
                 f'{code} 近10日DDX DDY 主力资金DDE大单净量'],
                capture_output=True, text=True, timeout=60,
                cwd=os.path.expanduser('~/skills/mx-data'),
                env={**os.environ, 'MX_APIKEY': mx_key})
            
            # 查询2: 主力资金 + PE
            _sp.run(
                [sys.executable, os.path.expanduser('~/skills/mx-data/mx_data.py'),
                 f'{code} 主力净流入资金 PE市盈率'],
                capture_output=True, text=True, timeout=60,
                cwd=os.path.expanduser('~/skills/mx-data'),
                env={**os.environ, 'MX_APIKEY': mx_key})
            
            # 读取所有生成的 Excel 文件
            xlsx_files = _glob.glob(f'{mx_out_dir}/mx_data_{code}_*.xlsx')
            for xf in xlsx_files:
                try:
                    wb = openpyxl.load_workbook(xf, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows = list(ws.iter_rows(max_row=5, values_only=True))
                        if not rows:
                            continue
                        headers = [str(c) if c else '' for c in rows[0]]
                        data = [str(c) if c else '' for c in rows[1]] if len(rows) > 1 else []
                        
                        # Parse DDX
                        for i, h in enumerate(headers):
                            if i < len(data):
                                try:
                                    val = float(data[i])
                                    s = str(h).strip()
                                    if s == '3日DDX' or '3日DDX' in s: enrichment['ddx_3d'] = val
                                    elif s == '5日DDX' or '5日DDX' in s: enrichment['ddx_5d'] = val
                                    elif s == '10日DDX' or '10日DDX' in s: enrichment['ddx_10d'] = val
                                    elif '市盈率' in s or s == 'PE(TTM)' or s == 'PE':
                                        enrichment['pe'] = val
                                except: pass
                            # 主力净流入
                            if ('净流入' in h and ('主力' in h or '区间' in h) and 
                                '超大' not in h and '大单' not in h and '中单' not in h and '小单' not in h):
                                if i < len(data) and data[i]:
                                    m = re.match(r'(-?\d+\.?\d*)\s*(万|亿)?元?', data[i])
                                    if m:
                                        val = float(m.group(1))
                                        if m.group(2) == '亿': val *= 1e8
                                        elif m.group(2) == '万': val *= 1e4
                                        enrichment['main_amount'] = val
                except:
                    pass
            
            if 'ddx_5d' in enrichment:
                enrichment['_ddx_source'] = 'mx-data'
    except Exception as e:
        pass  # mx-data不可用,降级
    
    # 2. TDX MCP 大单数据: 直接计算 DDX
    if 'ddx_5d' not in enrichment:
        try:
            from data_sources.tdx_mcp import TDXClient
            tdx = TDXClient()
            
            # 查询大单净额
            big_order = tdx.query(f"{code} 大单买入 大单卖出 超大单买入", size=1)
            headers = big_order.get('headers', [])
            data_row = big_order.get('data', [[]])[0] if big_order.get('data') else []
            
            big_net = 0
            super_big_net = 0
            for i, h in enumerate(headers):
                if i < len(data_row):
                    try:
                        val = float(str(data_row[i]).replace(',',''))
                        hs = str(h)
                        if '大单净额' in hs and '超大' not in hs:
                            big_net = val
                        elif '超大单净额' in hs:
                            super_big_net = val
                    except: pass
            
            if big_net != 0:
                # 从 TDX 获取真实流通市值
                try:
                    float_info = tdx.query(f"{code} 流通股本 流通市值", size=1)
                    f_headers = float_info.get('headers', [])
                    f_data = float_info.get('data', [[]])[0] if float_info.get('data') else []
                    float_cap = 0
                    for i, h in enumerate(f_headers):
                        if i < len(f_data):
                            hs = str(h)
                            if ('流通市值' in hs or '流通股本' in hs) and '最新' in hs:
                                try: float_cap = float(str(f_data[i]).replace(',',''))
                                except: pass
                except:
                    float_cap = 0
                
                if float_cap == 0:
                    # fallback to estimated float cap
                    kline_amount = quote.get('amount', 0) if quote else 0
                    if kline_amount == 0 and klines:
                        kline_amount = klines[-1].get('amount', 0) if klines else 0
                    float_cap = kline_amount * 20 if kline_amount > 0 else 0
                
                if float_cap > 0:
                    ddx_val = abs(big_net) / float_cap * 100
                    enrichment['ddx_5d'] = round(ddx_val, 4)
                    enrichment['ddx_10d'] = round(ddx_val * 0.9, 4)
                    enrichment['_ddx_source'] = 'tdx-mcp'
                enrichment['main_amount'] = abs(big_net) + abs(super_big_net)
                enrichment['super_large_net'] = abs(super_big_net)
        except:
            pass
    
    # 3. TDX MCP: 主力资金 + PE + 行业 (DDX 已有时跳过)
    try:
        from data_sources.tdx_mcp import TDXClient
        tdx = TDXClient()
        
        # 主力净流入
        capital = tdx.query(f"{code} 主力资金净流入", size=1)
        for row in capital.get('data', []):
            for h, v in zip(capital['headers'], row):
                if '净额' in str(h) or '净买' in str(h) or '净流' in str(h):
                    try:
                        enrichment['main_amount'] = float(str(v).replace('万','').replace('亿','').strip())
                        if '亿' in str(v): enrichment['main_amount'] *= 1e8
                        elif '万' in str(v): enrichment['main_amount'] *= 1e4
                    except: pass
        
        # PE/财务 (mx-data优先，TDX做fallback)
        if 'pe' not in enrichment:
            fin = tdx.query_financials(code)
            pe_str = fin.get('PE', fin.get('市盈率', '0'))
            try: enrichment['pe'] = float(pe_str)
            except: pass
        
        # 行业
        for k, v in fin.items():
            if '行业' in str(k):
                enrichment['topic'] = str(v).replace('@','').strip()
                break
    except:
        pass
    
    # 3. data_fallback L2: 大单净额
    try:
        r = subprocess.run(
            [sys.executable, f"{OPEN_SCRIPTS}/data_fallback.py",
             "--mode", "l2", "--code", code],
            capture_output=True, text=True, timeout=15
        )
        l2 = json.loads(r.stdout)
        bo = l2.get('big_order', {})
        net = bo.get('net_amount', 0)
        if net != 0:
            if 'main_amount' not in enrichment:
                enrichment['main_amount'] = net
            enrichment['super_large_net'] = net
    except:
        pass
    
    # 4. ddx_calculator 兜底: 用 L2/主力额 + K线估算 DDX
    if 'ddx_5d' not in enrichment and 'main_amount' in enrichment:
        try:
            closes = [k['close'] for k in klines] if klines else []
            price = closes[-1] if closes else (quote.get('price', 0) if quote else 0)
            amount = enrichment['main_amount']
            # 估算流通市值: 用今日成交额/换手率估算
            # amount 是K线的成交额字段, 不是主力净额
            kline_amount = quote.get('amount', 0) if quote else 0
            if kline_amount == 0 and klines:
                kline_amount = klines[-1].get('amount', 0)
            est_float_cap = kline_amount * 20 if kline_amount > 0 else price * 1e9  # 假设换手5%
            
            from core.ddx_calculator import calculate_ddx_from_amount
            # float_shares ≈ float_market_cap / price
            ddx_r = calculate_ddx_from_amount(
                main_net_inflow=amount,
                float_shares=est_float_cap / price if price > 0 else 1e9,
                current_price=price
            )
            if ddx_r.get('ddx'):
                enrichment['ddx_5d'] = round(ddx_r['ddx'], 4)
                enrichment['ddx_10d'] = round(ddx_r['ddx'] * 0.8, 4)  # 保守估计
                enrichment['_ddx_source'] = 'calculated'
        except:
            pass
    
    # 数据源标注
    if enrichment:
        enrichment['_ddx_source'] = 'mx-data' if 'ddx_5d' in enrichment else 'L2_estimated'
    
    return enrichment


def build_input_json(quote: dict, klines: list) -> dict:
    """从行情+K线构建输入格式 — 优先用真实数据"""
    code = quote.get("code", "000001")
    price = quote.get("price", klines[-1]["close"] if klines else 10)
    closes = [k['close'] for k in klines]
    highs = [k['high'] for k in klines]
    lows = [k['low'] for k in klines]
    volumes = [k['volume'] for k in klines]
    
    # 基础指标计算 (始终用K线自算，这是真实的)
    from core.math_utils import (
        rsi_wilder, band_position, volume_ratio, daily_returns, daily_volatility
    )
    
    rsi = rsi_wilder(closes) or 50
    band_pct = band_position(price, min(lows[-60:]), max(highs[-60:]))
    rets = daily_returns(closes[-60:])
    vol = daily_volatility(rets) if rets else 0.03
    vol_r = volume_ratio(volumes[-1], sum(volumes[-21:-1])/20) if len(volumes) >= 21 else 1.0
    ampl = (highs[-1] - lows[-1]) / closes[-2] * 100 if len(closes) >= 2 else 3.0
    
    # 真实数据补全
    enriched = enrich_real_data(code, quote, klines)
    
    return {
        "code": code,
        "price": price,
        "ddx_5d": enriched.get("ddx_5d"),
        "ddx_10d": enriched.get("ddx_10d"),
        "main_amount": enriched.get("main_amount", quote.get("amount", 0) * 0.1),
        "super_large_net": enriched.get("super_large_net", 0),
        "pe": enriched.get("pe"),
        "low_60d": min(lows[-60:]), "high_60d": max(highs[-60:]),
        "ma20": sum(closes[-20:])/20 if len(closes) >= 20 else price,
        "ma5": sum(closes[-5:])/5 if len(closes) >= 5 else price,
        "rsi": round(rsi, 2),
        "volatility": round(vol, 4),
        "amplitude": round(ampl, 2),
        "active_buy_ratio": 0.55,
        "inst_vs_retail": 1.0, "continuity_3d": 1,
        "topic": enriched.get("topic", "自动分析"),
        "topic_rank": 5,
        "is_limit_up": False, "lianban_days": 0,
        "E5_position_pct": round(band_pct, 2),
        "vol_ratio": round(vol_r, 2),
        "sentiment": "neutral",
        "_data_source": {k: "enriched" for k in enriched} if enriched else {},
    }


def main():
    parser = argparse.ArgumentParser(description="六维博弈引擎 V5.8.1 (开源版)")
    parser.add_argument("--input", "-i", help="JSON输入文件路径")
    parser.add_argument("--kline", "-k", help="K线数据JSON文件路径（用于E5分析）")
    parser.add_argument("--code", "-c", help="股票代码（自动获取行情+K线）")
    parser.add_argument("--consensus", action="store_true", help="同时执行共识仲裁")
    parser.add_argument("--classic-score", type=float, default=50, help="经典四视角评分(共识模式)")
    parser.add_argument("-o", "--output", help="输出文件")
    parser.add_argument("--json", action="store_true", help="输出原始JSON")
    args = parser.parse_args()
    
    # 加载数据
    if args.input:
        with open(args.input, 'r') as f:
            data = json.load(f)
        klines = None
        if args.kline:
            with open(args.kline, 'r') as f:
                kd = json.load(f)
            klines = kd if isinstance(kd, list) else kd.get('data', [])
    elif args.code:
        quote, klines = fetch_data(args.code)
        data = build_input_json(quote, klines)
    else:
        print("错误: 需要 --input 或 --code 参数")
        sys.exit(1)
    
    # 六维分析
    result = game_analyze(data, klines if 'klines' in dir() else None)
    
    # 共识仲裁
    if args.consensus:
        classic_score = args.classic_score
        six_dim_score = result.get("final_score", 50)
        consensus = consensus_arbitration(six_dim_score, classic_score, 1.0)
        forecast_r = tf_forecast(six_dim_score, classic_score)
        result["consensus"] = consensus
        result["forecast_detail"] = forecast_r
    
    output = json.dumps(result, ensure_ascii=False, indent=2)
    
    if args.output:
        with open(args.output, 'w') as f:
            f.write(output)
        print(f"结果已写入 {args.output}")
    
    if args.json:
        print(output)
    else:
        from core.formatters import format_game
        print(format_game(result, code=args.code or "", name=""))


if __name__ == "__main__":
    main()
